"""Backlog Mutabakat — UAT board ↔ TRADE/OPS board karşılaştırma motoru.

Amaç: Product'ın UAT board'unda (MBSUATEAM) açtığı taskların, çalışma
board'larındaki (MBSTRADE / MBSOPS) karşılıklarını bulmak; hangi UAT maddesinin
işleme alınmadığını (açıkta kalan iş) ve hangi hedef taskın kaynağının UAT'de
bulunmadığını ortaya çıkarmak. Analist tek tık çalıştırır, sonucu ekranda görür,
çok sayfalı Excel raporu indirir.

TASARIM İLKELERİ:
- **0 LLM tokenı.** Tamamen DETERMİNİSTİK — `claude -p`/MCP ÇAĞRILMAZ. Sadece Jira
  REST okuması + openpyxl yazımı. Saniyeler sürer, kullanıcı kotası harcanmaz.
- **Excel girişi YOK.** Eski sürüm elle yüklenen takip Excel'ini senkronlardı; bu
  sürüm iki board'u doğrudan Jira'dan çekip karşılaştırır. Çıktı, sıfırdan üretilen
  temiz bir mutabakat raporudur (cerrahi zip yazımına gerek yok — yeni dosya).

Eşleştirme (katmanlı, güçlüden zayıfa):
  1. KESİN   — UAT ile hedef arasında zaten issue-link (Jira bağlantısı) var.
  2. YÜKSEK  — link yok ama başlık+içerik Jaccard benzerliği ESIK_YUKSEK üstünde.
  3. ADAY    — Jaccard ESIK_ORTA ile ESIK_YUKSEK arasında → analist teyit eder.
  4. EŞLEŞMEYEN — hiçbiri tutmadı.

Tarama modu HEDEF board'a uygulanır (tüm board / epic-story altı / anahtar kelime);
UAT board'u her zaman TAM taranır (product oradan iş açar).
"""
from __future__ import annotations

import logging
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .atlassian import atlassian_post, jira_site_url
from .jira_gorevleri import (
    _benzerlik_jetonlari,
    _cloud_id,
    _iptal_statusu_mu,
    _ISSUE_ALANLARI,
    _issue_ayrıstir,
    _KAPSAYICI_TIP_ADLARI,
    alt_gorevleri_cek,
)

logger = logging.getLogger("analyst.backlog_mutabakat")

# ─── Sabitler ──────────────────────────────────────────────────────────────────
VARSAYILAN_UAT = "MBSUATEAM"
VARSAYILAN_HEDEF = ["MBSTRADE", "MBSOPS"]

ESIK_YUKSEK = 0.55   # bu skorun üstü otomatik EŞLEŞTİ
ESIK_ORTA = 0.35     # bu ile ESIK_YUKSEK arası → teyit bekleyen ADAY
_MIN_JETON = 4       # bu kadar token yoksa benzerlik sinyali güvenilmez

# UAT board'unda bu durumdaki tasklar kapsam dışı (hatalı/iptal kayıtlar; işleme alınmaz).
# JQL'de baştan elenir; ayrıca güvenlik ağı olarak çekilen kayıtlarda da filtrelenir.
# Not: tablodaki "≠" statünün parçası değil, UAT≠Hedef fark işaretidir.
# Yazım varyantları birlikte (Jira board'unda "Created in Error" görülüyor).
UAT_HARIC_DURUMLAR = ["Created in Error", "Create In Error"]

_ID_DESENI = re.compile(r"^[A-Z][A-Z0-9]+-\d+$")
_PROJE_DESENI = re.compile(r"^[A-Z][A-Z0-9]+$")


# ─── JQL yardımcıları ────────────────────────────────────────────────────────────
def _jql_ara(jql: str, cloud_id: str) -> list[dict]:
    """Verilen JQL'i sayfalayarak çeker ve görev sözlüğü listesi döndürür."""
    gorevler, next_token = [], None
    while True:
        body = {"jql": jql, "fields": _ISSUE_ALANLARI, "maxResults": 100}
        if next_token:
            body["nextPageToken"] = next_token
        data = atlassian_post("/rest/api/3/search/jql", body=body, cloud_id=cloud_id)
        for issue in data.get("issues", []):
            gorevler.append(_issue_ayrıstir(issue))
        next_token = data.get("nextPageToken")
        if not next_token:
            break
    return gorevler


def _keyleri_cek(keyler: list[str], cloud_id: str) -> list[dict]:
    """Verilen issue key'leri bulkfetch ile çeker ve parse eder.
    Taranan sette OLMAYAN ama UAT'a linkli hedef task'ları eşleştirmeye dahil etmek için
    kullanılır. Hata/bulunamama sessizce atlanır."""
    out: list[dict] = []
    for i in range(0, len(keyler), 100):   # bulkfetch tek istekte en fazla 100 key
        parca = keyler[i:i + 100]
        try:
            data = atlassian_post("/rest/api/3/issue/bulkfetch",
                                  body={"issueIdsOrKeys": parca, "fields": _ISSUE_ALANLARI},
                                  cloud_id=cloud_id)
        except Exception:
            logger.warning("Linkli hedef task'lar çekilemedi (atlanıyor): %s", parca)
            continue
        for issue in data.get("issues", []):
            out.append(_issue_ayrıstir(issue))
    return out


def _jql_kacis(deger: str) -> str:
    """JQL string literal içinde çift tırnak ve ters bölü kaçışlar."""
    return deger.replace("\\", "\\\\").replace('"', '\\"')


def _proje_listesi_jql(projeler: list[str]) -> str:
    icerik = ", ".join(f'"{_jql_kacis(p)}"' for p in projeler)
    return f"project in ({icerik})"


def _hedef_gorevleri_topla(mod: str, hedef_projeler: list[str],
                           hedef_keys: list[str], anahtar_kelime: str,
                           cloud_id: str) -> list[dict]:
    """Seçilen tarama moduna göre hedef board görevlerini toplar (tekrarsız).

    Epic modunda kapsam iki filtreyle daraltılır: girilen epic/story anahtarlarının
    ALTINDAKİ görevler + bunların YALNIZCA hedef board'lara (`hedef_projeler`) ait
    olanları. Böylece tüm board yerine sadece ilgili alt küme çekilir (token/zaman)."""
    if mod == "epic":
        birlesik: dict[str, dict] = {}
        for key in hedef_keys:
            for g in alt_gorevleri_cek(key):
                gk = g.get("key", "")
                proje = gk.split("-", 1)[0]
                if hedef_projeler and proje not in hedef_projeler:
                    continue   # epic altındaki ama hedef board dışı görevleri ele
                birlesik.setdefault(gk, g)
        return list(birlesik.values())

    jql = _proje_listesi_jql(hedef_projeler)
    if mod == "keyword":
        jql += f' AND text ~ "{_jql_kacis(anahtar_kelime)}"'
    jql += " ORDER BY created ASC"
    return _jql_ara(jql, cloud_id)


# ─── Benzerlik ──────────────────────────────────────────────────────────────────
def _jaccard(a: set[str], b: set[str]) -> float:
    if len(a) < _MIN_JETON or len(b) < _MIN_JETON:
        return 0.0
    birlesim = a | b
    if not birlesim:
        return 0.0
    return len(a & b) / len(birlesim)


def _sira_no(key: str) -> int:
    """Jira key'inin sonundaki sayı = board'daki sıra no (MBSUATEAM-116 → 116).
    Sayı yoksa listede sona düşsün diye çok büyük döner."""
    m = re.search(r"(\d+)$", key or "")
    return int(m.group(1)) if m else 10**9


def _kapsayici_tip_mi(g: dict) -> bool:
    """Görev Epic/Story gibi kapsayıcı bir tip mi? (parser 'type' = issuetype adı)."""
    return str(g.get("type", "")).strip().lower() in _KAPSAYICI_TIP_ADLARI


# İlişki metni "bağlılık" (dependency) mı yoksa gevşek "ilişki" mi? Jira link tip adları
# İngilizce/Türkçe olabilir; anahtar kelimeyle kaba sınıflama yapılır.
_BAGLILIK_IPUCLARI = ("block", "engel", "depend", "bağıml", "bagiml", "clone", "klon",
                      "duplicat", "mükerrer", "mukerrer", "cause", "neden", "split")


def _iliski_sinifi(iliski: str) -> str:
    """KESİN eşleşmedeki Jira link'inin türü: 'bağlılık' (blocks/depends/…) veya 'ilişki' (relates)."""
    t = (iliski or "").strip().casefold()
    return "bağlılık" if any(ip in t for ip in _BAGLILIK_IPUCLARI) else "ilişki"


# Story KÖPRÜSÜ için: yalnızca Story/Hikaye seviyesinde köprü kurulur. Epic/Initiative
# bilinçli olarak DIŞARIDA — onlar çok geniş kapsar, dolaylı eşleşmede yanlış pozitif üretir.
_KOPRU_TIP_ADLARI = {"story", "hikaye"}


def _kopru_link_mi(lk: dict) -> bool:
    """Bir issue-link'in hedefi Story/Hikaye mi? (transitif köprü adayı)."""
    return str(lk.get("type", "")).strip().lower() in _KOPRU_TIP_ADLARI


def _iptal_ayir(gorevler: list[dict]) -> tuple[list[dict], list[dict]]:
    """İptal edilmiş görevleri (İptal Edildi / CANCEL / CANCELED) ayırır.
    (iptaller, kalan) döndürür — böylece iptaller ana akışa girmez, kendi kovasında görünür."""
    iptal = [g for g in gorevler if _iptal_statusu_mu(g.get("status", ""))]
    kalan = [g for g in gorevler if not _iptal_statusu_mu(g.get("status", ""))]
    return iptal, kalan


def _sade(g: dict, proje_ekle: bool = False) -> dict:
    """Görev sözlüğünü ekran/rapor için sade kayda indirger."""
    key = g.get("key", "")
    kayit = {
        "sira": _sira_no(key),
        "key": key,
        "ozet": g.get("summary", ""),
        "durum": g.get("status", ""),
        "atanan": g.get("assignee", ""),
        "tur": g.get("type", ""),
    }
    if proje_ekle:
        kayit["proje"] = (key.split("-", 1)[0]) if key else ""
    return kayit


def _satir(uat: dict, hedef: dict, guven: str, gerekce: str, skor: float) -> dict:
    return {
        "sira": _sira_no(uat.get("key", "")),   # UAT board sıra no (sıralama + gösterim)
        "uat_key": uat.get("key", ""),
        "uat_ozet": uat.get("summary", ""),
        "uat_durum": uat.get("status", ""),
        "uat_atanan": uat.get("assignee", ""),
        "hedef_key": hedef.get("key", ""),
        "hedef_ozet": hedef.get("summary", ""),
        "hedef_durum": hedef.get("status", ""),
        "hedef_atanan": hedef.get("assignee", ""),
        "guven": guven,
        "gerekce": gerekce,
        "skor": round(skor, 2),
    }


# ─── Ana giriş ──────────────────────────────────────────────────────────────────
def mutabakat(uat_proje: str = VARSAYILAN_UAT,
              hedef_projeler: list[str] | None = None,
              mod: str = "tum",
              hedef_keys: list[str] | None = None,
              anahtar_kelime: str = "") -> dict:
    """UAT board'u ile hedef board(lar)ı karşılaştırır; eşleşen/eşleşmeyen kovalarını
    döndürür. LLM ÇAĞIRMAZ.

    mod: 'tum' (tüm hedef board), 'epic' (hedef_keys epic/story altı),
         'keyword' (hedef board'da anahtar_kelime geçen tasklar).
    """
    uat_proje = (uat_proje or VARSAYILAN_UAT).strip().upper()
    hedef_projeler = [p.strip().upper() for p in (hedef_projeler or VARSAYILAN_HEDEF) if p.strip()]
    hedef_keys = [k.strip().upper() for k in (hedef_keys or []) if k.strip()]
    anahtar_kelime = (anahtar_kelime or "").strip()
    mod = mod if mod in ("tum", "epic", "keyword") else "tum"

    if not _PROJE_DESENI.match(uat_proje):
        raise ValueError(f"Geçersiz UAT proje anahtarı: '{uat_proje}'")
    if not hedef_projeler:
        # Epic modunda da gerekli: epic altındaki görevleri bu board'larla sınırlarız.
        raise ValueError("En az bir hedef board proje anahtarı girin.")
    for p in hedef_projeler:
        if not _PROJE_DESENI.match(p):
            raise ValueError(f"Geçersiz hedef proje anahtarı: '{p}'")
    if mod == "epic":
        if not hedef_keys:
            raise ValueError("Epic/Story modunda en az bir epic/story anahtarı girin (örn. MBSTRADE-12).")
        for k in hedef_keys:
            if not _ID_DESENI.match(k):
                raise ValueError(f"Geçersiz epic/story anahtarı: '{k}' (örn. MBSTRADE-12 olmalı).")
    if mod == "keyword" and not anahtar_kelime:
        raise ValueError("Anahtar kelime modunda bir arama terimi girin.")

    cloud_id = _cloud_id()

    # ── Jira'dan çek ──
    # UAT board'unda "Create In Error" (ve UAT_HARIC_DURUMLAR'daki) tasklar kapsam dışı.
    uat_jql = f'project = "{_jql_kacis(uat_proje)}"'
    if UAT_HARIC_DURUMLAR:
        haric = ", ".join(f'"{_jql_kacis(d)}"' for d in UAT_HARIC_DURUMLAR)
        uat_jql += f" AND status NOT IN ({haric})"
    uat_jql += " ORDER BY created ASC"
    uat_gorevler = _jql_ara(uat_jql, cloud_id)
    # Güvenlik ağı: JQL durum adını eşleştiremezse (özel workflow) elde de filtrele.
    _haric_norm = {d.strip().casefold() for d in UAT_HARIC_DURUMLAR}
    uat_gorevler = [g for g in uat_gorevler
                    if (g.get("status") or "").strip().casefold() not in _haric_norm]
    hedef_gorevler = _hedef_gorevleri_topla(mod, hedef_projeler, hedef_keys, anahtar_kelime, cloud_id)

    # ── Kapsam dışı tipler: Epic/Story (kapsayıcı) elenir — yalnızca yaprak iş kalemleri karşılaştırılır.
    uat_gorevler   = [g for g in uat_gorevler   if not _kapsayici_tip_mi(g)]
    hedef_gorevler = [g for g in hedef_gorevler if not _kapsayici_tip_mi(g)]

    # ── İptal edilmiş task'lar ana akıştan ayrılır → kendi kovası ("İptal Edilenler").
    # Böylece eşleşen/açıkta kalan listeleriyle karışmaz; iptaller ayrı görünür.
    iptal_uat, uat_gorevler     = _iptal_ayir(uat_gorevler)
    iptal_hedef, hedef_gorevler = _iptal_ayir(hedef_gorevler)
    iptaller = [_sade(g, proje_ekle=True) for g in (iptal_uat + iptal_hedef)]
    iptaller.sort(key=lambda r: (r.get("proje", ""), r["sira"]))

    hedef_index = {g["key"]: g for g in hedef_gorevler}
    uat_index = {g["key"]: g for g in uat_gorevler}
    hedef_jeton = {g["key"]: _benzerlik_jetonlari(g) for g in hedef_gorevler}

    # ── Kapsam dışı ama LİNKLİ hedef task'ları tamamla ──
    # Bir UAT task, hedef proje(ler)deki bir key'e Jira ile bağlıysa ama o hedef task
    # taranan sette değilse (örn. epic/keyword modunda ya da alt-görev), yalnızca o linkli
    # hedef task'ları tek tek çekip eşleştirmeye dahil ederiz. Böylece gerçek ilişki
    # varken UAT "açıkta kalan iş" sanılmaz. Bu ek hedefler similarity/eşleşmeyen_hedef'e
    # KATILMAZ (yalnız KESİN link eşleşmesi için); iptal/Epic-Story olanlar dahil edilmez.
    _hedef_proje_set = set(hedef_projeler)
    _eksik_link_keyleri = {
        lk.get("key", "") for u in uat_gorevler for lk in u.get("baglantililar", [])
        if lk.get("key") and lk["key"] not in hedef_index
        and lk["key"].split("-", 1)[0] in _hedef_proje_set
    }
    link_hedef_index: dict[str, dict] = {}
    if _eksik_link_keyleri:
        for g in _keyleri_cek(sorted(_eksik_link_keyleri), cloud_id):
            if _kapsayici_tip_mi(g) or _iptal_statusu_mu(g.get("status", "")):
                continue   # Epic/Story ve iptal hedefler eşleşmeye alınmaz (ana akışla tutarlı)
            link_hedef_index[g["key"]] = g
        if link_hedef_index:
            logger.info("Kapsam dışı ama linkli %d hedef task eşleştirmeye dahil edildi: %s",
                        len(link_hedef_index), ", ".join(sorted(link_hedef_index)))

    def _hedef_bul(key: str) -> dict | None:
        return hedef_index.get(key) or link_hedef_index.get(key)

    eslesenler: list[dict] = []
    adaylar: list[dict] = []
    eslesen_uat: set[str] = set()
    aday_uat: set[str] = set()
    kullanilan_hedef: set[str] = set()   # eşleşen VEYA aday olarak bir UAT'a bağlanan hedefler

    # ── 1. KESİN: mevcut Jira bağlantıları ──
    # (uat_key, hedef_key) → hazır gerekçe metni. setdefault: bir çift zaten kurulduysa
    # ilk (daha güçlü/doğrudan) gerekçe korunur.
    kesin_ciftler: dict[tuple[str, str], str] = {}

    def _kesin_ekle(uk: str, hk: str, gerekce: str) -> None:
        kesin_ciftler.setdefault((uk, hk), gerekce)

    # 1a. DOĞRUDAN link: UAT↔hedef (kapsam dışı hedef dahil) + hedef tarafındaki linkler.
    for u in uat_gorevler:
        for lk in u.get("baglantililar", []):
            hk = lk.get("key", "")
            if hk and (hk in hedef_index or hk in link_hedef_index):
                iliski = lk.get("iliski", "ilişkili")
                notu = "" if hk in hedef_index else " · kapsam dışı hedef"
                _kesin_ekle(u["key"], hk, f"Jira {_iliski_sinifi(iliski)}: {u['key']} “{iliski}” {hk}{notu}")
    for h in hedef_gorevler:
        for lk in h.get("baglantililar", []):
            uk = lk.get("key", "")
            if uk in uat_index:
                iliski = lk.get("iliski", "ilişkili")
                _kesin_ekle(uk, h["key"], f"Jira {_iliski_sinifi(iliski)}: {h['key']} “{iliski}” {uk}")

    # 1b. STORY KÖPRÜSÜ (transitif): UAT task ve hedef task AYNI Story'ye linkliyse dolaylı eşleşir.
    # Story bir UAT taskıyla eşleşmişse, o story'ye bağlı hedef task'lar da eşleşmiş sayılır
    # (iş zaten üst/story üzerinden takip ediliyor). Epic değil, yalnızca Story seviyesinde köprü.
    uat_koprusu: dict[str, set[str]] = defaultdict(set)   # story_key → {uat_key}
    for u in uat_gorevler:
        for lk in u.get("baglantililar", []):
            if _kopru_link_mi(lk) and lk.get("key"):
                uat_koprusu[lk["key"]].add(u["key"])
    hedef_koprusu: dict[str, set[str]] = defaultdict(set)  # story_key → {hedef_key}
    for h in hedef_gorevler:
        for lk in h.get("baglantililar", []):
            if _kopru_link_mi(lk) and lk.get("key"):
                hedef_koprusu[lk["key"]].add(h["key"])
    kopru_sayisi = 0
    for skey, us in uat_koprusu.items():
        hs = hedef_koprusu.get(skey)
        if not hs:
            continue
        for uk in sorted(us):
            for hk in sorted(hs):
                if (uk, hk) not in kesin_ciftler:
                    kopru_sayisi += 1
                _kesin_ekle(uk, hk, f"Story köprüsü: {uk} ve {hk} ortak “{skey}” story’sine bağlı")
    if kopru_sayisi:
        logger.info("Story köprüsüyle %d dolaylı eşleşme eklendi.", kopru_sayisi)

    # ── Satırları kur ──
    for (uk, hk), gerekce in sorted(kesin_ciftler.items()):
        hedef_g = _hedef_bul(hk)
        if hedef_g is None:
            continue   # güvenlik: hedef verisi yoksa satır kurulamaz
        eslesenler.append(_satir(uat_index[uk], hedef_g, "Kesin", gerekce, 1.0))
        eslesen_uat.add(uk)
        kullanilan_hedef.add(hk)

    # ── 2/3. YÜKSEK / ADAY: içerik+başlık benzerliği (linki olmayan UAT'lar için) ──
    for u in uat_gorevler:
        if u["key"] in eslesen_uat:
            continue
        u_jeton = _benzerlik_jetonlari(u)
        en_iyi_hk, en_iyi_skor = None, 0.0
        for hk, hjeton in hedef_jeton.items():
            skor = _jaccard(u_jeton, hjeton)
            if skor > en_iyi_skor:
                en_iyi_hk, en_iyi_skor = hk, skor
        if en_iyi_hk is None:
            continue
        yuzde = round(en_iyi_skor * 100)
        if en_iyi_skor >= ESIK_YUKSEK:
            eslesenler.append(_satir(u, hedef_index[en_iyi_hk],
                                     "Yüksek", f"İçerik benzerliği %{yuzde}", en_iyi_skor))
            eslesen_uat.add(u["key"])
            kullanilan_hedef.add(en_iyi_hk)
        elif en_iyi_skor >= ESIK_ORTA:
            adaylar.append(_satir(u, hedef_index[en_iyi_hk],
                                  "Aday", f"İçerik benzerliği %{yuzde} — teyit bekliyor", en_iyi_skor))
            aday_uat.add(u["key"])
            kullanilan_hedef.add(en_iyi_hk)

    # ── 4. EŞLEŞMEYEN ──
    eslesmeyen_uat = [_sade(u) for u in uat_gorevler
                      if u["key"] not in eslesen_uat and u["key"] not in aday_uat]
    eslesmeyen_hedef = [_sade(h, proje_ekle=True) for h in hedef_gorevler
                        if h["key"] not in kullanilan_hedef]

    # Tüm kovalar UAT sıra no'suna (hedef kovası kendi key no'suna) göre artan sıralı —
    # karışık liste yerine 1, 2, 3… gitsin (analist hiçbir maddeyi atlamasın).
    eslesenler.sort(key=lambda r: r["sira"])
    adaylar.sort(key=lambda r: r["sira"])
    eslesmeyen_uat.sort(key=lambda r: r["sira"])
    eslesmeyen_hedef.sort(key=lambda r: r["sira"])

    logger.info("Mutabakat: UAT=%d hedef=%d → eşleşen=%d aday=%d eşleşmeyen_uat=%d eşleşmeyen_hedef=%d iptal=%d",
                len(uat_gorevler), len(hedef_gorevler), len(eslesenler), len(adaylar),
                len(eslesmeyen_uat), len(eslesmeyen_hedef), len(iptaller))

    return {
        "ok": True,
        "uat_proje": uat_proje,
        "hedef_projeler": hedef_projeler,
        "mod": mod,
        "jira_url": jira_site_url(cloud_id),
        "eslesenler": eslesenler,
        "adaylar": adaylar,
        "eslesmeyen_uat": eslesmeyen_uat,
        "eslesmeyen_hedef": eslesmeyen_hedef,
        "iptaller": iptaller,
        "sayimlar": {
            "uat_toplam": len(uat_gorevler),
            "hedef_toplam": len(hedef_gorevler),
            "eslesen": len(eslesenler),
            "aday": len(adaylar),
            "eslesmeyen_uat": len(eslesmeyen_uat),
            "eslesmeyen_hedef": len(eslesmeyen_hedef),
            "iptal": len(iptaller),
        },
    }


# ─── Excel raporu (openpyxl, sıfırdan) ───────────────────────────────────────────
_BASLIK_DOLGU = PatternFill("solid", fgColor="1F2937")
_BASLIK_YAZI = Font(bold=True, color="FFFFFF")
_SARILI = Alignment(wrap_text=True, vertical="top")


def _sayfa_yaz(ws, basliklar: list[str], satirlar: list[list], genislikler: list[int]) -> None:
    ws.append(basliklar)
    for c in range(1, len(basliklar) + 1):
        h = ws.cell(1, c)
        h.fill = _BASLIK_DOLGU
        h.font = _BASLIK_YAZI
        ws.column_dimensions[get_column_letter(c)].width = genislikler[c - 1]
    for satir in satirlar:
        ws.append(satir)
        for c in range(1, len(basliklar) + 1):
            ws.cell(ws.max_row, c).alignment = _SARILI
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(basliklar))}{max(ws.max_row, 1)}"


def rapor_uret(sonuc: dict, cikti_dir: str | Path) -> Path:
    """Mutabakat sonucundan çok sayfalı .xlsx raporu üretir; dosya yolunu döndürür."""
    cikti_dir = Path(cikti_dir)
    cikti_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Eşleşenler"
    esles_bas = ["Sıra", "UAT Key", "UAT Özet", "UAT Durum", "UAT Atanan", "Hedef Key", "Hedef Özet",
                 "Hedef Durum", "Hedef Atanan", "Eşleşme", "Gerekçe", "Skor"]
    esles_kaynak = sorted(sonuc.get("eslesenler", []) + sonuc.get("adaylar", []),
                          key=lambda r: r.get("sira", 10**9))
    esles_satir = [[r.get("sira", ""), r["uat_key"], r["uat_ozet"], r["uat_durum"],
                    r.get("uat_atanan", ""), r["hedef_key"],
                    r["hedef_ozet"], r["hedef_durum"], r.get("hedef_atanan", ""),
                    ("Aday" if r["guven"] == "Aday" else "Evet"), r["gerekce"], r["skor"]]
                   for r in esles_kaynak]
    _sayfa_yaz(ws1, esles_bas, esles_satir, [6, 14, 44, 14, 18, 14, 44, 14, 18, 10, 26, 8])

    ws2 = wb.create_sheet("Eşleşmeyen UAT")
    _sayfa_yaz(ws2, ["Sıra", "UAT Key", "Özet", "Durum", "Atanan", "Tür"],
               [[r.get("sira", ""), r["key"], r["ozet"], r["durum"], r.get("atanan", ""), r["tur"]]
                for r in sonuc.get("eslesmeyen_uat", [])],
               [6, 14, 52, 16, 18, 14])

    ws3 = wb.create_sheet("Eşleşmeyen TRADE-OPS")
    _sayfa_yaz(ws3, ["Sıra", "Hedef Key", "Proje", "Özet", "Durum", "Atanan", "Tür"],
               [[r.get("sira", ""), r["key"], r.get("proje", ""), r["ozet"], r["durum"],
                 r.get("atanan", ""), r["tur"]]
                for r in sonuc.get("eslesmeyen_hedef", [])],
               [6, 14, 12, 52, 16, 18, 14])

    ws4 = wb.create_sheet("İptal Edilenler")
    _sayfa_yaz(ws4, ["Sıra", "Key", "Proje", "Özet", "Durum", "Atanan", "Tür"],
               [[r.get("sira", ""), r["key"], r.get("proje", ""), r["ozet"], r["durum"],
                 r.get("atanan", ""), r["tur"]]
                for r in sonuc.get("iptaller", [])],
               [6, 14, 12, 52, 16, 18, 14])

    damga = datetime.now().strftime("%Y-%m-%d_%H%M")
    yol = cikti_dir / f"UAT_Mutabakat_{damga}.xlsx"
    wb.save(str(yol))
    logger.info("Mutabakat raporu üretildi: %s", yol.name)
    return yol
